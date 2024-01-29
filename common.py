from shared.config import Config
from shared.di import *
from datetime import datetime
import openpyxl
import pandas as pd

import os
from matplotlib import pyplot as plt

# 100% width!
from IPython.core.display import display, HTML

from shared.schema import TableSchema, TableImportSchema

display(HTML("<style>.container { width:100% !important; }</style>"))

# Float formatting
pd.options.display.float_format = '{:,.2f}'.format


def dump(df, tag="export", folderPath=r'c:\var', asHtml=False):
    import os
    if not os.path.exists(folderPath):
        os.makedirs(folderPath)

    if type(tag) == dict:
        tag = [[k, tag[k]] for k in tag]
        tag = makeFilename(tag)
    elif type(tag) == list:
        tag = makeFilename(tag)
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    basePath = os.path.join(folderPath, ts + "-" + tag)
    if (asHtml):
        f = open(basePath + '.html', 'w')
        f.write(df.to_html())
    else:  # assume is a dataframe
        writer = pd.ExcelWriter(
            basePath + ".xlsx",
            engine="xlsxwriter",
            datetime_format="yyyy-mm-dd hh:mm:ss",
            date_format="yyyy-mm-dd",
        )

        import pandas.io.formats.excel
        pandas.io.formats.excel.ExcelFormatter.header_style = None

        df.to_excel(writer, sheet_name="DATA", index=False)
        worksheet = writer.sheets['DATA']

        # workbook  = writer.book
        # header_format = workbook.add_format({
        #     'bold': False,
        #     'text_wrap': False,
        #     'valign': 'bottom'})

        # worksheet.set_row(0, None, header_format)
        worksheet.set_zoom(80)
        worksheet.autofit()

        writer.close()

        # wb = Workbook()
        # if ( isinstance(df, pd.io.formats.style.Styler) ):
        #     df = df.data
        # values = [df.columns] + list(df.values)
        # ws = wb.new_sheet("DATA", data=values)

        # wb.save( basePath + ".xlsx")


def dumpPlot(tag="export"):
    if type(tag) == list:
        tag = makeFilename(tag)
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    plt.savefig(os.path.join(r'c:\var', ts + "-" + tag + ".png"), dpi=600)


def makeFilename(attributes):
    return ''.join([f'({i[0]}{"@" + str(i[1]) if (len(i) > 1 and i[1] is not None) else ""})' for i in attributes])


class ConfigHelper():
    def __init__(self):
        self.cfg = injector.get_instance(Config)

    def getConfig(self):
        return self.cfg


def readTable(path, schema=None, sep=',', encoding=None, sheet_name="DEFAULT", date_parser=None):
    """Reads a table from CSV or Excel file and implies the file format from the file extension

    :param path: file path
    :param schema: schema (list of list) including columns to be read, types, and new column if not None, i.e. [ORIGINAL_COLUMN_NAME, TYPE, NEW_COLUMN_NAME_OR_NONE]
    :param sep: optional separator character, for CSV only
    :param encoding: optional encoding, for CSV only
    """

    import os
    import ast

    _, fileExtension = os.path.splitext(path)

    if (schema is not None):

        from datetime import datetime

        def mydateparser(numpy_array):
            return pd.to_datetime(pd.Series(numpy_array), format="%d/%m/%Y", errors='coerce')

        if date_parser is None:
            date_parser = mydateparser

        # [[import_name, type, new_name]]
        if (type(schema) == TableSchema):  # covert from legacy list of list to ImportSchema
            schema = [[f, f.type, f] for f in schema.fields]
        # elif (type(schema) == TableImportSchema):
        elif ('TableImportSchema' in str(type(schema))): # @hack workaround for type check failing for some reason
            schema = [[schema.importFields[i], f.type, f]
                      for i, f in enumerate(schema.fields)]
        else:
            schema = [[f[0], f[1], f[2]] if len(f) > 2 else [f[0], f[1]] for f in
                      schema]  # convert Field instance to string for input and target field names

        columnsDtype = {i[0]: (str if i[1] == datetime else i[1])
                        for i in schema}
        columnWhere = [i[0] for i in schema]
        columnRename = {i[0]: (i[2] if len(i) > 2 else i[0]) or i[0] for i in
                        schema}  # allow new name to be either not specified len(i[2]) == 2, or None
        columnDates = [i[0] for i in schema if i[1] == datetime]

        if fileExtension.startswith('.xl'):
            # df = pd.read_excel(path, dtype=columnsDtype, usecols=columnWhere, engine="openpyxl", parse_dates=columnDates, date_parser=date_parser)
            EXCEL_CSV_CONVERT_FROM_SIZE = 30 * 2 ** 20  # 30MB
            if os.path.getsize(path) > EXCEL_CSV_CONVERT_FROM_SIZE:
                generateExcelToCsv()
                from subprocess import check_output
                csv_path = path + '.csv'
                print(check_output(
                    ['cscript.exe', EXCELTOCSV_FILE, path, csv_path, sheet_name]))
                # pd.read_csv(csv)
                try:
                    df = pd.read_csv(csv_path, dtype=columnsDtype, usecols=columnWhere, sep=sep, encoding=encoding,
                                     parse_dates=columnDates, date_parser=date_parser)
                finally:
                    os.remove(csv_path)
            else:
                import warnings
                with warnings.catch_warnings(record=True):
                    warnings.simplefilter(
                        "always")  # silent warning "Workbook contains no default style, apply openpyxl's default "
                    df = pd.read_excel(path, dtype=columnsDtype, usecols=columnWhere, engine="openpyxl",
                                       parse_dates=columnDates, date_parser=date_parser)
        else:
            # df = pd.read_csv(path, dtype=columnsDtype, usecols=columnWhere, sep=sep, encoding=encoding)
            df = pd.read_csv(path, dtype=columnsDtype, usecols=columnWhere, sep=sep, encoding=encoding,
                             parse_dates=columnDates, date_parser=date_parser)

        for columnName, columnType in columnsDtype.items():
            if columnType == list:
                df[columnName] = df[columnName].apply(
                    lambda x: ast.literal_eval(x))  # convert list back to list type
            elif columnType == str:
                # for strings, replace NaN with empty string
                df[columnName] = df[columnName].fillna('')

        df.rename(columns=columnRename, inplace=True)
    else:
        if fileExtension.startswith('.xl'):
            df = pd.read_excel(path)
        else:
            df = pd.read_csv(path)

    return df


EXCELTOCSV_FILE = r'c:\var\ExcelToCsv.vbs'


def generateExcelToCsv():
    # write vbscript to file
    vbscript = """
    if WScript.Arguments.Count < 3 Then
        WScript.Echo "Please specify the source and the destination files. Usage: ExcelToCsv <xls/xlsx source file> <csv destination file> <worksheet number (starts at 1)>"
        Wscript.Quit
    End If

    csv_format = 6

    Set objFSO = CreateObject("Scripting.FileSystemObject")

    src_file = objFSO.GetAbsolutePathName(Wscript.Arguments.Item(0))
    dest_file = objFSO.GetAbsolutePathName(WScript.Arguments.Item(1))
    worksheet_name = WScript.Arguments.Item(2)
    If worksheet_name = "DEFAULT" Then
    worksheet_name = 1
    End If

    Dim oExcel
    Set oExcel = CreateObject("Excel.Application")
    On Error Resume Next
    Set oExcel = CreateObject("Excel.Application")
    If Err.Number <>0 Then
        WScript.Echo "Start Excel failed, try again"
        WScript.Sleep 100 'wait for 0.1 second
        Set oExcel = CreateObject("Excel.Application")
    End If
    On Error GoTo 0

    Dim oBook
    Set oBook = oExcel.Workbooks.Open(src_file)
    oBook.Worksheets(worksheet_name).Activate

    oBook.SaveAs dest_file, csv_format

    oBook.Close False
    oExcel.Quit
    """

    if not os.path.exists(EXCELTOCSV_FILE):
        f = open(EXCELTOCSV_FILE, 'w')
        # f.write(vbscript.encode('utf-8'))
        f.write(vbscript)
        f.close()


def remove_file(path):
    if os.path.exists(path):
        os.remove(path)


def read_excel(path, sheet_name="DEFAULT", remove_csv=True):
    generateExcelToCsv()
    from subprocess import check_output
    csv_path = path + '.csv'
    dtypes_path = path + '.dtypes.csv'
    if not os.path.exists(csv_path):
        print(check_output(
            ['cscript.exe', EXCELTOCSV_FILE, path, csv_path, sheet_name]))
        dtypes = get_meta(path)

        dtypes.to_frame().to_csv(dtypes_path)
    try:
        dtype = pd.read_csv(dtypes_path, index_col=0)
        dt_cols = list(dtype[dtype['0'].str.contains('datetime')].index)
        if len(dt_cols) > 0:
            df = pd.read_csv(csv_path, parse_dates=dt_cols)
        else:
            df = pd.read_csv(csv_path)

    finally:
        if remove_csv:
            remove_file(csv_path)
            remove_file(dtypes_path)
        # pass
    return df


def get_meta(path):
    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet = workbook.active
    columns = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    data = [list(row) for row in sheet.iter_rows(
        min_row=2, max_row=2, values_only=True)]
    workbook.close()
    df = pd.DataFrame(data, columns=columns)
    return df.dtypes
